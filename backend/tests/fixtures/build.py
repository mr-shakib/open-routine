"""Build synthetic routine workbooks for tests.

The real DIU workbook is not redistributable and its exact layout has not been
verified, so tests run against sheets we construct here in the layout the parser
documents. Every awkward case the parser must survive is represented: merged day
blocks, a lab spanning two slots, a TBA teacher, an elective, split lab
subsections and a blank cell.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

SLOT_HEADERS = [
    "08:30-10:00",
    "10:00-11:30",
    "11:30-01:00",
    "01:00-02:30",
    "02:30-04:00",
    "04:00-05:30",
]

#: (day, room, {slot_index: cell_text})
ROWS: list[tuple[str, str, dict[int, str]]] = [
    ("Saturday", "KT-503\n  (COM LAB)", {0: "CSE414(62_E1)\nSRH", 1: "CSE332(60_C)\nMAH"}),
    ("Saturday", "KT-515", {0: "CSE311(60_C)\nNRC", 3: "TCSE412(61_A)\nSAH"}),
    ("Saturday", "G1-007  (COM LAB)", {2: "CSE414(62_E2)\nSRH"}),
    ("Sunday", "KT-503\n  (COM LAB)", {1: "CSE414(62_E1)\nSRH"}),
    ("Sunday", "KT-515", {0: "CSE311(60_C)\nNRC", 1: "CSE322(60_C)\nTBA"}),
    ("Monday", "KT-801(A)", {4: "CSE311(60_C)\nMAH"}),
    ("Tuesday", "KT-515", {5: "CSE332(60_C)\nMAH"}),
    ("Wednesday", "G1-007  (COM LAB)", {0: "CSE414(62_E2)\nSRH"}),
    ("Thursday", "KT-503\n  (COM LAB)", {2: "CSE322(60_C)\nNRC"}),
]


def build_workbook(path: Path, *, merge_days: bool = True) -> Path:
    """Write a valid routine workbook and return its path."""
    workbook = Workbook()
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    worksheet.title = "Routine"

    worksheet.cell(row=1, column=1, value="Day")
    worksheet.cell(row=1, column=2, value="Room")
    for offset, header in enumerate(SLOT_HEADERS):
        worksheet.cell(row=1, column=3 + offset, value=header)

    for index, (day, room, cells) in enumerate(ROWS):
        row = 2 + index
        worksheet.cell(row=row, column=1, value=day)
        worksheet.cell(row=row, column=2, value=room)
        for slot_index, text in cells.items():
            worksheet.cell(row=row, column=3 + slot_index, value=text)

    if merge_days:
        # Day labels are merged down their block in the published sheet; the
        # reader must expand them or every row but the first loses its day.
        _merge_day_blocks(worksheet)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path


def _merge_day_blocks(worksheet: Worksheet) -> None:
    start = 2
    for index in range(1, len(ROWS) + 1):
        at_end = index == len(ROWS)
        changed = not at_end and ROWS[index][0] != ROWS[index - 1][0]
        if at_end or changed:
            end = 1 + index
            if end > start:
                worksheet.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
                for row in range(start + 1, end + 1):
                    worksheet.cell(row=row, column=1, value=None)
            start = end + 1


def build_broken_workbook(path: Path) -> Path:
    """A workbook whose slot headers do not match the lattice."""
    workbook = Workbook()
    worksheet = workbook.active
    assert isinstance(worksheet, Worksheet)
    worksheet.cell(row=1, column=1, value="Day")
    worksheet.cell(row=1, column=2, value="Room")
    for offset, header in enumerate(["09:00-10:00", "10:00-11:00", "11:00-12:00"]):
        worksheet.cell(row=1, column=3 + offset, value=header)
    worksheet.cell(row=2, column=1, value="Saturday")
    worksheet.cell(row=2, column=2, value="KT-101")
    worksheet.cell(row=2, column=3, value="CSE101(60_A)\nABC")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path
