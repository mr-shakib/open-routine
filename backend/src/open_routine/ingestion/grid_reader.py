"""Walk a routine spreadsheet and yield one raw record per populated cell.

Expected sheet layout -- a day-blocked grid::

                 | 08:30-10:00 | 10:00-11:30 | ... |   <- slot header row
    -------------+-------------+-------------+-----+
    Saturday | KT-503 | CSE414(62_E1) SRH |   |     |
             | KT-515 |             | CSE311(60_C) MAH |
    Sunday   | KT-503 | ...
    ---------+--------+

* one column holds the day (often merged down its block),
* one column holds the room,
* the remaining columns are the six time slots.

Both the day and room columns are located by inspection rather than assumed, so
minor layout drift between revisions does not break the import.

.. warning::
   The exact published layout has not been verified against a real DIU file.
   Header detection is heuristic and deliberately fails loudly via
   :func:`~open_routine.ingestion.lattice.validate_lattice` when the six slot
   columns cannot be found. Validate against a real routine before trusting an
   import.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from open_routine.core.errors import IngestionError
from open_routine.ingestion.lattice import normalise_day, normalise_slot, validate_lattice
from open_routine.ingestion.normalizer import normalise_room, normalise_whitespace


@dataclass(frozen=True, slots=True)
class RawCell:
    """One populated cell, before its text is parsed."""

    day: str
    time_slot: str
    room: str
    room_type: str
    text: str
    row: int
    column: int


@dataclass(frozen=True, slots=True)
class SheetLayout:
    """Where the axes live in a particular sheet."""

    header_row: int
    day_column: int
    room_column: int
    slot_columns: dict[int, str]  #: worksheet column index -> canonical slot


def read_grid(path: str | Path, *, sheet: str | None = None) -> list[RawCell]:
    """Read a routine workbook and return every populated cell."""
    workbook = load_workbook(filename=str(path), data_only=True, read_only=False)
    try:
        worksheet = workbook[sheet] if sheet else workbook.worksheets[0]
        return read_worksheet(worksheet)
    finally:
        workbook.close()


def read_worksheet(worksheet: Worksheet) -> list[RawCell]:
    grid = _materialise(worksheet)
    layout = _detect_layout(grid)

    days_seen: list[str] = []
    cells: list[RawCell] = []

    current_day: str | None = None
    current_room: str | None = None
    current_room_type = "Theory"

    for row_index in range(layout.header_row + 1, len(grid)):
        row = grid[row_index]

        day_value = _at(row, layout.day_column)
        if day_value:
            resolved = normalise_day(day_value)
            if resolved:
                current_day = resolved
                if resolved not in days_seen:
                    days_seen.append(resolved)

        room_value = _at(row, layout.room_column)
        if room_value:
            room, room_type = normalise_room(room_value)
            if room:
                current_room, current_room_type = room, room_type

        if not current_day or not current_room:
            continue

        for column_index, slot in layout.slot_columns.items():
            text = _at(row, column_index)
            if not text or not text.strip():
                continue
            cells.append(
                RawCell(
                    day=current_day,
                    time_slot=slot,
                    room=current_room,
                    room_type=current_room_type,
                    text=text,
                    row=row_index + 1,
                    column=column_index + 1,
                )
            )

    validate_lattice(days_seen, list(layout.slot_columns.values()))
    return cells


def _materialise(worksheet: Worksheet) -> list[list[str]]:
    """Read the sheet into a dense text grid, expanding merged ranges.

    A class spanning two slots, or a day label spanning its block of rows, is a
    single merged cell whose value openpyxl reports only at the top-left corner.
    Left unexpanded, every other covered cell reads as blank and those classes
    vanish silently -- which is the failure mode most likely to go unnoticed.
    """
    height = worksheet.max_row or 0
    width = worksheet.max_column or 0
    grid: list[list[str]] = [["" for _ in range(width)] for _ in range(height)]

    for row in worksheet.iter_rows(min_row=1, max_row=height, max_col=width):
        for cell in row:
            value = cell.value
            if value is None:
                continue
            grid[cell.row - 1][cell.column - 1] = _stringify(value)

    for merged in worksheet.merged_cells.ranges:
        top, left = merged.min_row - 1, merged.min_col - 1
        if top >= height or left >= width:
            continue
        value = grid[top][left]
        if not value:
            continue
        for r in range(merged.min_row - 1, min(merged.max_row, height)):
            for c in range(merged.min_col - 1, min(merged.max_col, width)):
                grid[r][c] = value

    return grid


def _stringify(value: Any) -> str:
    if isinstance(value, str):
        return value
    return str(value)


def _at(row: list[str], index: int) -> str:
    return row[index] if 0 <= index < len(row) else ""


def _detect_layout(grid: list[list[str]]) -> SheetLayout:
    """Locate the header row and the day/room/slot columns."""
    best: SheetLayout | None = None
    best_score = 0

    for row_index, row in enumerate(grid[:40]):  # headers live near the top
        slot_columns: dict[int, str] = {}
        for column_index, value in enumerate(row):
            slot = normalise_slot(normalise_whitespace(value))
            if slot and slot not in slot_columns.values():
                slot_columns[column_index] = slot
        if len(slot_columns) > best_score:
            first_slot_column = min(slot_columns)
            best_score = len(slot_columns)
            best = SheetLayout(
                header_row=row_index,
                day_column=_find_day_column(grid, row_index, first_slot_column),
                room_column=_find_room_column(grid, row_index, first_slot_column),
                slot_columns=slot_columns,
            )

    if best is None or best_score == 0:
        from open_routine.ingestion.lattice import SLOTS

        raise IngestionError(
            "Could not find the time-slot header row, so the sheet does not "
            "match the expected routine lattice.",
            detail={"expected_slots": list(SLOTS), "rows_examined": min(len(grid), 40)},
        )
    return best


def _find_day_column(grid: list[list[str]], header_row: int, before: int) -> int:
    """The column left of the slots holding the most recognisable day names."""
    counts = {
        column: sum(1 for row in grid[header_row + 1 :] if normalise_day(_at(row, column)))
        for column in range(before)
    }
    return max(counts, key=lambda c: counts[c]) if counts and max(counts.values()) else 0


def _find_room_column(grid: list[list[str]], header_row: int, before: int) -> int:
    """The remaining column left of the slots with the most distinct values."""
    day_column = _find_day_column(grid, header_row, before)
    counts = {
        column: len(
            {
                normalise_whitespace(_at(row, column))
                for row in grid[header_row + 1 :]
                if normalise_whitespace(_at(row, column))
            }
        )
        for column in range(before)
        if column != day_column
    }
    return max(counts, key=lambda c: counts[c]) if counts else max(before - 1, 0)
